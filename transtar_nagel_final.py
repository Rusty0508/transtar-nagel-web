#!/usr/bin/env python3
"""
TRANSTAR-NAGEL FINAL SYSTEM v4.0
Финальная версия с правильным сопоставлением по номерам заказов
и извлечением всех необходимых данных
"""

import os
import re
import sys
import hashlib
import logging
import pandas as pd
import pdfplumber
from pathlib import Path
from datetime import datetime
from dataclasses import dataclass, field
from typing import List, Dict, Optional, Tuple
from collections import defaultdict
import warnings
warnings.filterwarnings('ignore')

# ═══════════════════════════════════════════════════════════════════
#                         КОНФИГУРАЦИЯ
# ═══════════════════════════════════════════════════════════════════

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('logs/transtar_final.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger('TRANSTAR-FINAL')

# ═══════════════════════════════════════════════════════════════════
#                      СТРУКТУРЫ ДАННЫХ
# ═══════════════════════════════════════════════════════════════════

@dataclass
class TransportOrder:
    """Транспортный заказ с полной информацией"""
    # Основные поля
    order_number: str
    date: str
    vehicle: str
    tour: str = ""

    # Точки маршрута
    loading_points: List[str] = field(default_factory=list)
    unloading_points: List[str] = field(default_factory=list)

    # Километраж
    empty_km: int = 0
    loaded_km: int = 0
    total_km: int = 0
    gps_km: int = 0

    # Финансовые данные
    planned_freight: float = 0.0
    planned_maut: float = 0.0
    planned_total: float = 0.0

    # Данные из гутшрифта
    gutschrift_amount: float = 0.0
    gutschrift_freight: float = 0.0
    gutschrift_maut: float = 0.0
    gutschrift_number: str = ""
    gutschrift_date: str = ""

    # Процент оплаты
    payment_percent: int = 100

    # Служебные поля
    file_name: str = ""

    def __post_init__(self):
        """Вычисление производных полей"""
        self.total_km = self.empty_km + self.loaded_km
        self.planned_total = self.planned_freight + self.planned_maut
        # GPS km = total_km по умолчанию
        if not self.gps_km:
            self.gps_km = self.total_km

    def format_tour(self) -> str:
        """Форматирование маршрута"""
        if self.loading_points or self.unloading_points:
            # Форматируем как в эталонном CSV
            loads = [self._format_city(p) for p in self.loading_points] if self.loading_points else []
            unloads = [self._format_city(p) for p in self.unloading_points] if self.unloading_points else []

            # Убираем дубликаты, сохраняя порядок
            seen = set()
            unique_loads = []
            for l in loads:
                if l and l not in seen:
                    unique_loads.append(l)
                    seen.add(l)

            unique_unloads = []
            for u in unloads:
                if u and u not in seen:
                    unique_unloads.append(u)
                    seen.add(u)

            # Форматируем маршрут
            if unique_loads and unique_unloads:
                if len(unique_loads) == 1 and len(unique_unloads) == 1:
                    return f"{unique_loads[0]}-{unique_unloads[0]}"
                elif len(unique_loads) == 1 and len(unique_unloads) > 1:
                    return f"{unique_loads[0]}-" + "+".join(unique_unloads)
                else:
                    return "-".join(unique_loads + unique_unloads)
            elif unique_loads:
                return "-".join(unique_loads)
            elif unique_unloads:
                return "-".join(unique_unloads)
        return self.tour or ""

    def _format_city(self, address: str) -> str:
        """Извлечение города из адреса"""
        # Пытаемся найти город после индекса
        if match := re.search(r'D\s*\d{5}\s+([A-Z\-]+)', address):
            city = match.group(1)
            # Сокращаем длинные названия
            if "GROSS-GERAU" in city:
                return "Groß-Gerau"
            elif "ESCHWEILER" in city:
                return "Eschweiler"
            elif "VOELKLINGEN" in city or "VÖLKLINGEN" in city:
                return "Völklingen"
            elif "SAARLOUIS" in city:
                return "Saarlouis"
            elif "TROISDORF" in city:
                return "Troisdorf"
            elif "LADENBURG" in city:
                return "Ladenburg"
            elif "RASTATT" in city:
                return "Rastatt"
            elif "BEXBACH" in city:
                return "Bexbach"
            elif "WESEL" in city:
                return "Wesel"
            elif "LANGENFELD" in city:
                return "Langenfeld"
            elif "BOCHUM" in city:
                return "Bochum"
            elif "HELMOND" in city:
                return "Helmond"
            elif "ESSEN" in city:
                return "Essen"
            elif "KLEINBLITTERSDORF" in city:
                return "Kleinblittersdorf"
            elif "DORTMUND" in city:
                return "Dortmund"
            elif "WUNSTORF" in city:
                return "Wunstorf"
            elif "HEDDESHEIM" in city:
                return "Heddesheim"
            elif "RAUNHEIM" in city:
                return "Raunheim"
            elif "ROSBACH" in city:
                return "Rosbach"
            elif "HAUNECK" in city:
                return "Hauneck"
            elif "KÖLN" in city or "KOLN" in city:
                return "Köln"
            elif "LEVERKUSEN" in city:
                return "Leverkusen"
            else:
                return city.title()

        # Альтернативный метод - ищем Na, или другие компании
        if "NAGEL" in address:
            return "Na"
        elif "EDEKA" in address:
            return "Edeka"
        elif "BAKERMAN" in address:
            return "Bakerman"
        elif "LACTALIS" in address:
            return "Lactalis"

        return ""

@dataclass
class GutschriftDetail:
    """Детализация заказа в гутшрифте"""
    transport_order: str  # Номер заказа
    date: str
    vehicle: str
    freight: float
    maut: float
    total: float
    route: str = ""

@dataclass
class Gutschrift:
    """Гутшрифт с детализацией по заказам"""
    number: str
    date: str
    period_from: str
    period_to: str
    total_freight: float = 0.0
    total_maut: float = 0.0
    total_amount: float = 0.0
    vat_amount: float = 0.0
    gross_amount: float = 0.0
    order_count: int = 0
    file_name: str = ""

    # Детализация по заказам
    details: List[GutschriftDetail] = field(default_factory=list)

# ═══════════════════════════════════════════════════════════════════
#                      ПАРСЕР ДОКУМЕНТОВ
# ═══════════════════════════════════════════════════════════════════

class FinalDocumentParser:
    """Финальный парсер с полным извлечением данных"""

    def __init__(self):
        self.errors = []

    def parse_german_number(self, text: str) -> float:
        """Конвертация немецкого числа в float"""
        try:
            # Удаляем точки (разделители тысяч) и заменяем запятую на точку
            return float(text.replace('.', '').replace(',', '.'))
        except:
            return 0.0

    def parse_transport_order(self, pdf_path: Path) -> Optional[TransportOrder]:
        """Парсинг транспортного заказа"""
        try:
            with pdfplumber.open(pdf_path) as pdf:
                # Извлекаем весь текст
                full_text = ""
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        full_text += page_text + "\n"

                if not full_text:
                    logger.warning(f"Пустой PDF: {pdf_path.name}")
                    return None

                # Создаем объект заказа
                order = TransportOrder(
                    order_number="",
                    date="",
                    vehicle="",
                    file_name=pdf_path.name
                )

                # Извлекаем номер заказа
                if match := re.search(r'TRN-(\d{4}\s?\d{2})', full_text):
                    order.order_number = match.group(1).replace(' ', '')
                else:
                    logger.warning(f"Не найден номер заказа в {pdf_path.name}")
                    return None

                # Извлекаем дату
                if match := re.search(r'Datum:\s*(\d{2}\.\d{2}\.\d{4})', full_text):
                    order.date = match.group(1)

                # Извлекаем транспортное средство
                if match := re.search(r'LKW-Kennzeichen:\s*([A-Z\s\-]+\d+)', full_text):
                    order.vehicle = match.group(1).strip()

                # Извлекаем километраж
                if match := re.search(r'//(\d+)\s*LEERKM\s*//(\d+)\s*LAST\s*KM', full_text):
                    order.empty_km = int(match.group(1))
                    order.loaded_km = int(match.group(2))

                # Извлекаем процент оплаты
                if match := re.search(r'//\s*(\d+)%', full_text):
                    order.payment_percent = int(match.group(1))

                # Извлекаем финансовые данные
                # Ищем строку с ценами
                price_section = re.search(r'Frachtpreis:.*?Maut:.*?(\d+[,\.]\d+)\s*EUR\s+(\d+[,\.]\d+)\s*EUR', full_text, re.DOTALL)
                if price_section:
                    order.planned_freight = self.parse_german_number(price_section.group(1))
                    order.planned_maut = self.parse_german_number(price_section.group(2))

                # Извлекаем точки загрузки (Ladestellen)
                loading_section = re.search(r'Ladestellen[^:]*:\s*(.*?)(?=\(Die vorgegebenen|Ladung:|$)', full_text, re.DOTALL)
                if loading_section:
                    text = loading_section.group(1).strip()
                    # Разбиваем по строкам и ищем адреса
                    lines = text.split('\n')
                    current_address = []

                    for line in lines:
                        line = line.strip()
                        if not line or 'Die vorgegebenen' in line:
                            break
                        # Ищем полный адрес (компания, улица, индекс и город)
                        if re.search(r'[A-Z].*D\s*\d{5}', line):
                            # Это строка с адресом и индексом
                            # Извлекаем компанию, улицу, индекс и город
                            match = re.search(r'([A-Z][\w\s\-&\.]+?),\s*([^,]+),\s*D\s*(\d{5})', line)
                            if match:
                                company = match.group(1).strip()
                                street = match.group(2).strip()
                                plz = match.group(3)

                                # Ищем город после индекса (до следующих чисел или конца строки)
                                rest_text = line[match.end():]
                                city_match = re.search(r'^\s*([A-Z][A-Z\-]*)', rest_text)
                                city = city_match.group(1) if city_match else ""

                                full_address = f"{company}, {street}, D {plz} {city}".strip()
                                order.loading_points.append(full_address)

                # Извлекаем точки разгрузки (Empfänger)
                empfanger_section = re.search(r'Empfänger[^:]*:\s*(.*?)(?=Frachtpreis|Zahlungsziel|$)', full_text, re.DOTALL)
                if empfanger_section:
                    text = empfanger_section.group(1).strip()
                    # Разбиваем текст и ищем адреса
                    lines = text.split('\n')

                    for line in lines:
                        line = line.strip()
                        if not line:
                            continue
                        # Ищем полный адрес
                        if re.search(r'[A-Z].*D\s*\d{5}', line):
                            # Извлекаем компанию, улицу, индекс и город
                            match = re.search(r'([A-Z][\w\s\-&\.]+?),\s*([^,]+),\s*D\s*(\d{5})', line)
                            if match:
                                company = match.group(1).strip()
                                street = match.group(2).strip()
                                plz = match.group(3)

                                # Ищем город после индекса
                                rest_text = line[match.end():]
                                city_match = re.search(r'^\s*([A-Z][A-Z\-]*)', rest_text)
                                city = city_match.group(1) if city_match else ""

                                # Убираем "Buchungsnr.:" и другие служебные слова
                                if 'Buchungsnr' not in company and 'zeiten' not in company:
                                    full_address = f"{company}, {street}, D {plz} {city}".strip()
                                    order.unloading_points.append(full_address)

                # Альтернативный поиск - LEER IN для точки загрузки
                if not order.loading_points:
                    if match := re.search(r'LEER IN\s+([A-Z\-]+)', full_text):
                        order.loading_points = [match.group(1)]

                # Альтернативный поиск - Tourbeginn/Tourende в комментарии к километражу
                if not order.loading_points or not order.unloading_points:
                    # Ищем в инструкциях
                    instruction = re.search(r'LADEINSTRUKTIONEN:.*?(?=A\.|$)', full_text, re.DOTALL)
                    if instruction:
                        instr_text = instruction.group(0)
                        # Извлекаем города из инструкций
                        if 'LEER IN' in instr_text:
                            if match := re.search(r'LEER IN\s+([A-Z\-]+)', instr_text):
                                if not order.loading_points:
                                    order.loading_points = [match.group(1)]

                logger.info(f"Обработан заказ {order.order_number} от {order.date} ({order.vehicle})")
                return order

        except Exception as e:
            logger.error(f"Ошибка при парсинге {pdf_path.name}: {str(e)}")
            return None

    def parse_gutschrift(self, pdf_path: Path) -> Optional[Gutschrift]:
        """Парсинг гутшрифта с детализацией"""
        try:
            with pdfplumber.open(pdf_path) as pdf:
                # Извлекаем весь текст
                full_text = ""
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        full_text += page_text + "\n"

                if not full_text:
                    logger.warning(f"Пустой PDF: {pdf_path.name}")
                    return None

                # Создаем объект гутшрифта
                gutschrift = Gutschrift(
                    number="",
                    date="",
                    period_from="",
                    period_to="",
                    file_name=pdf_path.name
                )

                # Извлекаем номер
                if match := re.search(r'Nr\.:\s*(\d+)', full_text):
                    gutschrift.number = match.group(1)
                else:
                    logger.warning(f"Не найден номер гутшрифта в {pdf_path.name}")
                    return None

                # Извлекаем дату
                if match := re.search(r'vom:\s*(\d{2}\.\d{2}\.\d{4})', full_text):
                    gutschrift.date = match.group(1)

                # Извлекаем период
                if match := re.search(r'Leistungszeitraum:\s*(\d{2}\.\d{2}\.\d{4})\s*-\s*(\d{2}\.\d{2}\.\d{4})', full_text):
                    gutschrift.period_from = match.group(1)
                    gutschrift.period_to = match.group(2)

                # Извлекаем общие суммы с первой страницы
                if match := re.search(r'Fracht\s+ST\s+[\d,\.]+\s+([\d,\.]+)', full_text):
                    gutschrift.total_freight = self.parse_german_number(match.group(1))

                if match := re.search(r'Mautkosten.*?ST\s+[\d,\.]+\s+([\d,\.]+)', full_text):
                    gutschrift.total_maut = self.parse_german_number(match.group(1))

                if match := re.search(r'Gesamtbetrag:\s*([\d,\.]+)\s*EUR', full_text):
                    gutschrift.gross_amount = self.parse_german_number(match.group(1))

                # Извлекаем количество заказов
                if match := re.search(r'Anzahl.*?Transportaufträge.*?:\s*(\d+)', full_text):
                    gutschrift.order_count = int(match.group(1))

                # Парсим детализацию (обычно на второй странице)
                # Ищем блоки с Transp.A.
                detail_pattern = r'Transp\.A\.\s+.*?\n.*?\n(\d{6})\s+(\d{2}\.\d{2}\.\d{4}).*?([A-Z\s\-]+\d+).*?Fracht.*?D\s+[\d,]+\s+([\d,\.]+)\s+EUR.*?Mautkosten.*?D\s+[\d,]+\s+([\d,\.]+)\s+EUR.*?Summe\s+([\d,\.]+)\s+EUR'

                for match in re.finditer(detail_pattern, full_text, re.DOTALL):
                    detail = GutschriftDetail(
                        transport_order=match.group(1),
                        date=match.group(2),
                        vehicle=match.group(3).strip(),
                        freight=self.parse_german_number(match.group(4)),
                        maut=self.parse_german_number(match.group(5)),
                        total=self.parse_german_number(match.group(6))
                    )
                    gutschrift.details.append(detail)
                    logger.debug(f"  - Детализация: заказ {detail.transport_order}, машина {detail.vehicle}, сумма {detail.total}")

                logger.info(f"Обработан гутшрифт {gutschrift.number} от {gutschrift.date} с {len(gutschrift.details)} заказами")
                return gutschrift

        except Exception as e:
            logger.error(f"Ошибка при парсинге {pdf_path.name}: {str(e)}")
            import traceback
            traceback.print_exc()
            return None

# ═══════════════════════════════════════════════════════════════════
#                    ПРОЦЕССОР ДОКУМЕНТОВ
# ═══════════════════════════════════════════════════════════════════

class FinalDocumentProcessor:
    """Финальный процессор с правильным сопоставлением"""

    def __init__(self, base_path: str = "."):
        self.base_path = Path(base_path)
        self.parser = FinalDocumentParser()

        # Хранилища данных
        self.transport_orders: List[TransportOrder] = []
        self.gutschrifts: List[Gutschrift] = []
        self.matched_data: List[Dict] = []

        # Создание папок
        self._create_folder_structure()

    def _create_folder_structure(self):
        """Создание структуры папок"""
        folders = [
            self.base_path / "documents" / "orders",
            self.base_path / "documents" / "gutschrifts",
            self.base_path / "output",
            self.base_path / "logs"
        ]

        for folder in folders:
            folder.mkdir(parents=True, exist_ok=True)

    def load_documents(self):
        """Загрузка всех документов"""
        print("\n📂 Загрузка документов...")

        # Используем переменные класса или значения по умолчанию
        orders_path = self.orders_path if hasattr(self, 'orders_path') else self.base_path / "documents" / "orders"
        gutschrifts_path = self.gutschrifts_path if hasattr(self, 'gutschrifts_path') else self.base_path / "documents" / "gutschrifts"

        print(f"📁 Ищу заказы в: {orders_path}")
        print(f"📁 Ищу гутшрифты в: {gutschrifts_path}")

        # Загрузка транспортных заказов
        if orders_path.exists():
            for pdf_file in sorted(orders_path.glob("*.pdf")):
                print(f"   Обработка: {pdf_file.name}")
                order = self.parser.parse_transport_order(pdf_file)
                if order:
                    self.transport_orders.append(order)

        # Загрузка гутшрифтов
        if gutschrifts_path.exists():
            for pdf_file in sorted(gutschrifts_path.glob("*.pdf")):
                print(f"   Обработка: {pdf_file.name}")
                gutschrift = self.parser.parse_gutschrift(pdf_file)
                if gutschrift:
                    self.gutschrifts.append(gutschrift)

        print(f"✅ Загружено заказов: {len(self.transport_orders)}")
        print(f"✅ Загружено гутшрифтов: {len(self.gutschrifts)}")

    def match_documents(self):
        """Сопоставление документов по номерам заказов"""
        print("\n🔗 Сопоставление документов...")

        matched_count = 0

        # Создаем индекс детализации гутшрифтов
        detail_index = {}
        for gs in self.gutschrifts:
            for detail in gs.details:
                detail_index[detail.transport_order] = {
                    'gutschrift': gs,
                    'detail': detail
                }

        # Сопоставляем каждый заказ
        for order in self.transport_orders:
            if order.order_number in detail_index:
                info = detail_index[order.order_number]
                gutschrift = info['gutschrift']
                detail = info['detail']

                # Заполняем данные из гутшрифта
                order.gutschrift_number = gutschrift.number
                order.gutschrift_date = gutschrift.date
                order.gutschrift_freight = detail.freight
                order.gutschrift_maut = detail.maut

                # Если процент оплаты не 100%, корректируем сумму
                if order.payment_percent != 100:
                    order.gutschrift_amount = detail.total * (order.payment_percent / 100)
                else:
                    order.gutschrift_amount = detail.total

                matched_count += 1
                logger.info(f"✅ Сопоставлен заказ {order.order_number} с гутшрифтом {gutschrift.number}")
            else:
                logger.warning(f"⚠️ Не найден гутшрифт для заказа {order.order_number}")

        print(f"✅ Сопоставлено: {matched_count} из {len(self.transport_orders)} заказов")

    def generate_report(self) -> pd.DataFrame:
        """Генерация отчета в формате эталонного CSV"""
        print("\n📊 Генерация отчета...")

        # Подготовка данных для DataFrame
        data = []
        for order in self.transport_orders:
            # Форматируем маршрут
            tour = order.format_tour()

            # Расчет разницы
            price_difference = order.gutschrift_amount - order.planned_total if order.gutschrift_amount else 0
            km_difference = order.gps_km - order.total_km

            row = {
                'Datum': order.date,
                'Tournummer': order.order_number,
                'Tour': tour,
                'Ladestellen': ' | '.join(order.loading_points) if order.loading_points else "",
                'Entladestellen': ' | '.join(order.unloading_points) if order.unloading_points else "",
                'Auftrag_km': order.total_km,
                'GPS_km': order.gps_km,
                'Maut_Auftrag': f"{order.planned_maut:.2f} €",
                'Maut_Gefahren': f"{order.gutschrift_maut:.2f} €" if order.gutschrift_maut else "",
                'Preis_Plan': f"{order.planned_freight:.2f} €",
                'Preis_Tatsächlich': f"{order.planned_total:.2f} €",
                'Gutschrift': f"{order.gutschrift_amount:.2f} €" if order.gutschrift_amount else "",
                'Kosten_Auftrag': f"{order.planned_total:.2f} €",
                'Kosten_Gefahren': f"{order.gutschrift_amount:.2f} €" if order.gutschrift_amount else "",
                'Differenz_km': km_difference,
                'Differenz_Preis': f"{price_difference:.2f} €" if price_difference else "",
                'Bearbeiter': 'EV',
                'GS_Datum': f"{order.gutschrift_number}/{order.gutschrift_date}" if order.gutschrift_number else "",
                'LKW': order.vehicle,
                'Prozent': f"{order.payment_percent}%"
            }
            data.append(row)

        # Создание DataFrame
        df = pd.DataFrame(data)

        # Сортировка по дате только если есть данные
        if not df.empty and 'Datum' in df.columns:
            df['Datum_sort'] = pd.to_datetime(df['Datum'], format='%d.%m.%Y', errors='coerce')
            df = df.sort_values('Datum_sort')
            df = df.drop('Datum_sort', axis=1)

        # Добавление итогов только если есть данные
        if self.transport_orders:
            total_row = {
                'Datum': 'GESAMT',
                'Tournummer': '',
                'Tour': '',
                'Ladestellen': '',
                'Entladestellen': '',
                'Auftrag_km': sum(o.total_km for o in self.transport_orders),
                'GPS_km': sum(o.gps_km for o in self.transport_orders),
                'Maut_Auftrag': f"{sum(o.planned_maut for o in self.transport_orders):.2f} €",
                'Maut_Gefahren': f"{sum(o.gutschrift_maut for o in self.transport_orders):.2f} €",
                'Preis_Plan': f"{sum(o.planned_freight for o in self.transport_orders):.2f} €",
                'Preis_Tatsächlich': f"{sum(o.planned_total for o in self.transport_orders):.2f} €",
                'Gutschrift': f"{sum(o.gutschrift_amount for o in self.transport_orders):.2f} €",
                'Kosten_Auftrag': f"{sum(o.planned_total for o in self.transport_orders):.2f} €",
                'Kosten_Gefahren': f"{sum(o.gutschrift_amount for o in self.transport_orders):.2f} €",
                'Differenz_km': sum(o.gps_km - o.total_km for o in self.transport_orders),
                'Differenz_Preis': f"{sum(o.gutschrift_amount - o.planned_total for o in self.transport_orders if o.gutschrift_amount):.2f} €",
                'Bearbeiter': '',
                'GS_Datum': '',
                'LKW': '',
                'Prozent': ''
            }

            df = pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)

        return df

    def export_to_excel_custom(self, df: pd.DataFrame, output_file: str, stats: dict = None) -> str:
        """Экспорт в Excel с форматированием для веб-интерфейса"""
        from openpyxl.styles import PatternFill, Alignment, Font

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Основной отчет
            if df is not None and not df.empty:
                df.to_excel(writer, sheet_name='Hauptbericht', index=False)

                # Получаем лист для форматирования
                workbook = writer.book
                worksheet = writer.sheets['Hauptbericht']

                # Пастельно-красный цвет для проблемных ячеек
                red_fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
                # Пастельно-желтый для предупреждений
                yellow_fill = PatternFill(start_color='FFF9E6', end_color='FFF9E6', fill_type='solid')

                # Проходим по строкам данных (начиная со 2-й строки, т.к. 1-я - заголовки)
                for row_num in range(2, len(df) + 2):
                    # Получаем индекс строки в DataFrame
                    df_row_idx = row_num - 2

                    if df_row_idx < len(df):
                        row_data = df.iloc[df_row_idx]

                        # Проверяем проблемные условия:

                        # 1. Если заказ не сопоставлен с гутшрифтом (пустое поле GS_Datum)
                        if pd.isna(row_data.get('GS_Datum')) or row_data.get('GS_Datum') == '':
                            for col in range(1, worksheet.max_column + 1):
                                worksheet.cell(row=row_num, column=col).fill = yellow_fill

                        # 2. Если есть большая разница в километраже (>10%)
                        try:
                            if 'Auftrag_km' in row_data and 'GPS_km' in row_data:
                                auftrag_km = row_data['Auftrag_km']
                                gps_km = row_data['GPS_km']
                                if isinstance(auftrag_km, (int, float)) and isinstance(gps_km, (int, float)):
                                    if auftrag_km > 0 and abs(gps_km - auftrag_km) / auftrag_km > 0.1:
                                        # Выделяем колонки с километражом
                                        for col_name in ['Auftrag_km', 'GPS_km', 'Differenz_km']:
                                            if col_name in df.columns:
                                                col_idx = df.columns.get_loc(col_name) + 1
                                                worksheet.cell(row=row_num, column=col_idx).fill = red_fill
                        except:
                            pass

                        # 3. Если процент оплаты меньше 100%
                        try:
                            if 'Prozent' in row_data:
                                prozent_str = str(row_data['Prozent']).replace('%', '')
                                if prozent_str and prozent_str.replace('.', '').isdigit():
                                    prozent_val = float(prozent_str)
                                    if prozent_val < 100:
                                        # Выделяем колонку процента
                                        col_idx = df.columns.get_loc('Prozent') + 1
                                        worksheet.cell(row=row_num, column=col_idx).fill = yellow_fill
                        except:
                            pass

                        # 4. Если есть отрицательная разница в цене
                        try:
                            if 'Differenz_Preis' in row_data:
                                diff_preis = str(row_data['Differenz_Preis']).replace('€', '').replace(',', '.').strip()
                                if diff_preis and diff_preis.replace('.', '').replace('-', '').isdigit():
                                    if float(diff_preis) < -50:  # Если разница больше -50 евро
                                        col_idx = df.columns.get_loc('Differenz_Preis') + 1
                                        worksheet.cell(row=row_num, column=col_idx).fill = red_fill
                        except:
                            pass

                # Автоматическая ширина колонок
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

            # Лист Gutschriften - сводка по гутшрифтам
            gs_data = []
            for gs in self.gutschrifts:
                gs_data.append({
                    'Nummer': gs.number,
                    'Datum': gs.date,
                    'Periode': f"{gs.period_from} - {gs.period_to}",
                    'Fracht': f"{gs.total_freight:.2f} €",
                    'Maut': f"{gs.total_maut:.2f} €",
                    'Gesamt': f"{gs.gross_amount:.2f} €",
                    'Anzahl_Aufträge': gs.order_count,
                    'Details': len(gs.details)
                })

            if gs_data:
                gs_df = pd.DataFrame(gs_data)
                gs_df.to_excel(writer, sheet_name='Gutschriften', index=False)

            # Лист GS_Details - детализация гутшрифтов
            detail_data = []
            for gs in self.gutschrifts:
                for detail in gs.details:
                    detail_data.append({
                        'GS_Nummer': gs.number,
                        'Transport_Auftrag': detail.transport_order,
                        'Datum': detail.date,
                        'LKW': detail.vehicle,
                        'Fracht': f"{detail.freight:.2f} €",
                        'Maut': f"{detail.maut:.2f} €",
                        'Summe': f"{detail.total:.2f} €"
                    })

            if detail_data:
                detail_df = pd.DataFrame(detail_data)
                detail_df.to_excel(writer, sheet_name='GS_Details', index=False)

            # Лист Statistik - статистика
            if stats:
                matched_orders = [o for o in self.transport_orders if o.gutschrift_number]
                unmatched_orders = [o for o in self.transport_orders if not o.gutschrift_number]

                stats_data = {
                    'Метрика': [
                        'Всего заказов',
                        'Всего гутшрифтов',
                        'Сопоставлено заказов',
                        'Не сопоставлено',
                        'Общий километраж (план)',
                        'Общий километраж (факт)',
                        'Общий фрахт (план)',
                        'Общий маут (план)',
                        'Итоговая сумма (план)',
                        'Итоговая сумма (гутшрифт)',
                        'Общая разница'
                    ],
                    'Значение': [
                        len(self.transport_orders),
                        len(self.gutschrifts),
                        len(matched_orders),
                        len(unmatched_orders),
                        sum(o.total_km for o in self.transport_orders) if self.transport_orders else 0,
                        sum(o.gps_km for o in self.transport_orders) if self.transport_orders else 0,
                        f"{sum(o.planned_freight for o in self.transport_orders):.2f} €" if self.transport_orders else "0.00 €",
                        f"{sum(o.planned_maut for o in self.transport_orders):.2f} €" if self.transport_orders else "0.00 €",
                        f"{sum(o.planned_total for o in self.transport_orders):.2f} €" if self.transport_orders else "0.00 €",
                        f"{sum(o.gutschrift_amount for o in matched_orders):.2f} €" if matched_orders else "0.00 €",
                        f"{sum(o.gutschrift_amount - o.planned_total for o in matched_orders):.2f} €" if matched_orders else "0.00 €"
                    ]
                }
                stats_df = pd.DataFrame(stats_data)
                stats_df.to_excel(writer, sheet_name='Statistik', index=False)

            # Лист Nicht_zugeordnet - несопоставленные заказы
            unmatched_orders = [o for o in self.transport_orders if not o.gutschrift_number]
            if unmatched_orders:
                unmatched_data = []
                for order in unmatched_orders:
                    unmatched_data.append({
                        'Tournummer': order.order_number,
                        'Datum': order.date,
                        'LKW': order.vehicle,
                        'Tour': order.format_tour(),
                        'Summe': f"{order.planned_total:.2f} €"
                    })
                unmatched_df = pd.DataFrame(unmatched_data)
                unmatched_df.to_excel(writer, sheet_name='Nicht_zugeordnet', index=False)

        return output_file

    def export_to_excel(self, df: pd.DataFrame) -> str:
        """Экспорт в Excel с форматированием"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = self.base_path / "output" / f"transtar_final_{timestamp}.xlsx"

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Основной отчет
            df.to_excel(writer, sheet_name='Hauptbericht', index=False)

            # Детали гутшрифтов
            gs_data = []
            for gs in self.gutschrifts:
                gs_data.append({
                    'Nummer': gs.number,
                    'Datum': gs.date,
                    'Periode': f"{gs.period_from} - {gs.period_to}",
                    'Fracht': f"{gs.total_freight:.2f} €",
                    'Maut': f"{gs.total_maut:.2f} €",
                    'Gesamt': f"{gs.gross_amount:.2f} €",
                    'Anzahl_Aufträge': gs.order_count,
                    'Details': len(gs.details)
                })

            if gs_data:
                gs_df = pd.DataFrame(gs_data)
                gs_df.to_excel(writer, sheet_name='Gutschriften', index=False)

            # Детализация гутшрифтов
            detail_data = []
            for gs in self.gutschrifts:
                for detail in gs.details:
                    detail_data.append({
                        'GS_Nummer': gs.number,
                        'Transport_Auftrag': detail.transport_order,
                        'Datum': detail.date,
                        'LKW': detail.vehicle,
                        'Fracht': f"{detail.freight:.2f} €",
                        'Maut': f"{detail.maut:.2f} €",
                        'Summe': f"{detail.total:.2f} €"
                    })

            if detail_data:
                detail_df = pd.DataFrame(detail_data)
                detail_df.to_excel(writer, sheet_name='GS_Details', index=False)

            # Статистика
            matched_orders = [o for o in self.transport_orders if o.gutschrift_number]
            unmatched_orders = [o for o in self.transport_orders if not o.gutschrift_number]

            stats_data = {
                'Метрика': [
                    'Всего заказов',
                    'Всего гутшрифтов',
                    'Сопоставлено заказов',
                    'Не сопоставлено',
                    'Общий километраж (план)',
                    'Общий километраж (факт)',
                    'Общий фрахт (план)',
                    'Общий маут (план)',
                    'Итоговая сумма (план)',
                    'Итоговая сумма (гутшрифт)',
                    'Общая разница'
                ],
                'Значение': [
                    len(self.transport_orders),
                    len(self.gutschrifts),
                    len(matched_orders),
                    len(unmatched_orders),
                    sum(o.total_km for o in self.transport_orders),
                    sum(o.gps_km for o in self.transport_orders),
                    f"{sum(o.planned_freight for o in self.transport_orders):.2f} €",
                    f"{sum(o.planned_maut for o in self.transport_orders):.2f} €",
                    f"{sum(o.planned_total for o in self.transport_orders):.2f} €",
                    f"{sum(o.gutschrift_amount for o in matched_orders):.2f} €",
                    f"{sum(o.gutschrift_amount - o.planned_total for o in matched_orders):.2f} €"
                ]
            }
            stats_df = pd.DataFrame(stats_data)
            stats_df.to_excel(writer, sheet_name='Statistik', index=False)

            # Несопоставленные заказы
            if unmatched_orders:
                unmatched_data = []
                for order in unmatched_orders:
                    unmatched_data.append({
                        'Tournummer': order.order_number,
                        'Datum': order.date,
                        'LKW': order.vehicle,
                        'Tour': order.format_tour(),
                        'Summe': f"{order.planned_total:.2f} €"
                    })
                unmatched_df = pd.DataFrame(unmatched_data)
                unmatched_df.to_excel(writer, sheet_name='Nicht_zugeordnet', index=False)

        print(f"✅ Отчет сохранен: {output_file}")
        return str(output_file)

# ═══════════════════════════════════════════════════════════════════
#                         ГЛАВНАЯ ФУНКЦИЯ
# ═══════════════════════════════════════════════════════════════════

def main():
    """Главная функция программы"""
    print("\n" + "═" * 60)
    print(" TRANSTAR-NAGEL FINAL SYSTEM v4.0 ".center(60))
    print(" Полное сопоставление по номерам заказов ".center(60))
    print("═" * 60)

    try:
        # Инициализация процессора
        processor = FinalDocumentProcessor()

        # Загрузка документов
        processor.load_documents()

        if not processor.transport_orders:
            print("❌ Транспортные заказы не найдены!")
            return None

        if not processor.gutschrifts:
            print("⚠️ Гутшрифты не найдены! Продолжаем без сопоставления...")

        # Сопоставление документов
        processor.match_documents()

        # Генерация отчета
        df = processor.generate_report()

        # Экспорт в Excel
        output_file = processor.export_to_excel(df)

        # Вывод результатов
        print("\n" + "🎉" * 30)
        print("\n✨ ОБРАБОТКА ЗАВЕРШЕНА УСПЕШНО!")
        print(f"\n📊 Результаты:")
        print(f"├── Обработано заказов: {len(processor.transport_orders)}")
        print(f"├── Обработано гутшрифтов: {len(processor.gutschrifts)}")
        matched = sum(1 for o in processor.transport_orders if o.gutschrift_number)
        print(f"├── Сопоставлено: {matched}")
        print(f"├── Не сопоставлено: {len(processor.transport_orders) - matched}")
        print(f"└── Файл отчета: {output_file}")

        # Вывод несопоставленных заказов
        unmatched = [o for o in processor.transport_orders if not o.gutschrift_number]
        if unmatched:
            print(f"\n⚠️ Не сопоставленные заказы:")
            for order in unmatched[:5]:  # Первые 5
                print(f"   - {order.order_number} от {order.date}")
            if len(unmatched) > 5:
                print(f"   ... и еще {len(unmatched) - 5}")

        print("\n" + "🎉" * 30)

        return processor, df

    except Exception as e:
        logger.error(f"Критическая ошибка: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

if __name__ == "__main__":
    main()